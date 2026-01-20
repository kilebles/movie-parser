from collections.abc import Sequence
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

from src.entities import Movie


class ExcelExporter:
    def export(self, movies: Sequence[Movie], output_path: Path) -> None:
        """Export movies to an Excel file."""
        wb = Workbook()
        ws = wb.active
        ws.title = "Top Movies"

        headers = ["#", "Название", "Год", "Рейтинг", "Ссылка"]
        header_font = Font(bold=True)

        for col, header in enumerate(headers, start=1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center")

        for row, movie in enumerate(movies, start=2):
            ws.cell(row=row, column=1, value=movie.rank)
            ws.cell(row=row, column=2, value=movie.title)
            ws.cell(row=row, column=3, value=movie.year)
            ws.cell(row=row, column=4, value=movie.rating)

            link_cell = ws.cell(row=row, column=5, value=movie.title)
            link_cell.hyperlink = movie.url
            link_cell.font = Font(color="0563C1", underline="single")

        column_widths = [5, 40, 8, 10, 40]
        for col, width in enumerate(column_widths, start=1):
            ws.column_dimensions[get_column_letter(col)].width = width

        wb.save(output_path)
